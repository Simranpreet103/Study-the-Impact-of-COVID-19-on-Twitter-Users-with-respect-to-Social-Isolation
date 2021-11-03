from textblob import TextBlob
import sys
import csv
import xlrd
from xlwt import Workbook
from nltk.corpus import wordnet
from os.path import join
from jsonpath_ng import jsonpath, parse
import openpyxl
from pathlib import Path
import xlsxwriter
import pandas as pd
from openpyxl import load_workbook
from textblob import TextBlob

xlsx_file = Path('Result2.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)
wsheet = wb_obj.active

fileName = "Result3.xlsx"
workbook = xlsxwriter.Workbook(fileName)
worksheet = workbook.add_worksheet("DataSet")

rowNum = -1
col = 0
for row in wsheet.iter_rows(max_row=wsheet.max_row):
    for cell in row:
        tweet = cell.value
        col = 1
        rowNum = rowNum + 1
        worksheet.write(rowNum, 0, tweet)
        print(tweet)
        j= TextBlob(tweet)
        t= j.sentiment.polarity
        #if t >= 0.1:
         #   k="positive"
        #elif t <= -0.1:
         #   k="negative"
        #else:
         #   k="neutral"
        worksheet.write(rowNum,1,t)
        #rowNum = rowNum + 1
workbook.close()
