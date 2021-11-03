import json
import os
from os.path import join
from ibm_watson import ToneAnalyzerV3
from ibm_watson.tone_analyzer_v3 import ToneInput
from ibm_cloud_sdk_core.authenticators import IAMAuthenticator
from jsonpath_ng import jsonpath, parse
import openpyxl
from pathlib import Path
import xlsxwriter

# Authentication via IAM
authenticator = IAMAuthenticator('*******************') # your api authentication
service = ToneAnalyzerV3(version='2017-09-21', authenticator=authenticator)
service.set_service_url('https://api.us-south.tone-analyzer.watson.cloud.ibm.com/instances/************************') #your_link

def analyseTone(tweet, row):
    print(tweet)
    tone = json.dumps(service.tone(tone_input=tweet,content_type="text/plain").get_result(),indent=2)
    # tone = '{   "document_tone": {     "tones": [       {         "score": 0.961678,         "tone_id": "joy",         "tone_name": "Joy"       },       {         "score": 0.983213,         "tone_id": "confident",         "tone_name": "Confident"       }     ]   },   "sentences_tone": [     {       "sentence_id": 0,       "text": "I am very happy.",       "tones": [         {           "score": 1.0,           "tone_id": "joy",           "tone_name": "Joy"         },         {           "score": 0.97759,           "tone_id": "confident",           "tone_name": "Confident"         }       ]     },     {       "sentence_id": 1,       "text": "It is a good day.",       "tones": [         {           "score": 0.914543,           "tone_id": "joy",           "tone_name": "Joy"         }       ]     }   ] }'
    Data = json.loads(tone)
    docTone = Data["document_tone"]

    col = 1
    row = row + 1
    worksheet.write(row, 0, tweet)
    for temp in docTone:
        jTone = parse("$.tones[*].tone_name");
        jScore = parse("$.tones[*].score")
        tone = jTone.find(docTone)
        score = jScore.find(docTone)
        i=0
        for tones in tone:
            print(tones.value)
            print(score[i].value)
            worksheet.write(row, col, tones.value)
            col = col+1
            worksheet.write(row, col, score[i].value)
            col = col+1
            i= i+1


#opening the tweet file
xlsx_file = Path('tweets.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)
wsheet = wb_obj.active


fileName = "ToneAnalysed_TweetDataSet.xlsx"
workbook = xlsxwriter.Workbook(fileName)
worksheet = workbook.add_worksheet("DataSet")

rowNum = -1
col = 0
for row in wsheet.iter_rows(max_row=wsheet.max_row):
    for cell in row:
        tweet = cell.value
        try:
            analyseTone(tweet, rowNum)
        except:
            print("An exception occurred")
        rowNum = rowNum + 1
workbook.close()
